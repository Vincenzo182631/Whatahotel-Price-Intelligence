# -*- coding: utf-8 -*-
"""Build the rice credit ledger workbook from the customer sales & loan ledger screenshot."""
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

OUT = "/home/user/Whatahotel-Price-Intelligence/ledger/Rice_Credit_Ledger.xlsx"

FONT = "Arial"
NAVY   = "1F3864"
HEADBG = "1F3864"
YELLOW = "FFFF00"
RED    = "FF0000"
GREY   = "D9D9D9"
BAND   = "F2F5FA"
GREEN  = "006100"

thin = Side(style="thin", color="B7BFCC")
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

# (customer, contact, date, product, qty, unit_price, method, flag)
ROWS = [
    ("Lorna",                 "", "6/8/2026",  "Rice (50 kg/sack)",  1, 2800, "CREDIT", ""),
    ("Ate Lenie",             "", "6/8/2026",  "Rice (50 kg/sack)",  1, 2800, "CREDIT", ""),
    ("Mingo",                 "", "6/15/2026", "Rice (25 kg/sack)",  2, 1400, "CREDIT", ""),
    ("Julius",                "", "6/16/2026", "Rice (25 kg/sack)",  1, 1400, "CREDIT", ""),
    ("Junard",                "", "6/16/2026", "Rice (25 kg/sack)",  1, 1400, "CREDIT", ""),
    ("Juaton Crossing",       "", "6/18/2026", "Rice (25 kg/sack)",  2, 1400, "CREDIT", ""),
    ("Nanay Juaton",          "", "6/18/2026", "Rice (25 kg/sack)",  2, 1300, "CASH",   ""),
    ("Catmon Tinadahan",      "", "6/20/2026", "Rice (25 kg/sack)",  1, 1300, "CASH",   ""),
    ("Junjun SM",             "", "6/20/2026", "Rice (25 kg/sack)",  1, 1400, "CREDIT", ""),
    ("Tita Susan",            "", "6/21/2026", "Rice (25 kg/sack)",  1, 1400, "CREDIT", "RED"),
    ("Juaton Crossing",       "", "6/20/2026", "Rice (50 kg/sack)",  2, 2800, "CREDIT", "RED"),
    ("Ranny",                 "", "6/20/2026", "Rice (50 kg/sack)",  1, 2800, "CREDIT", "RED"),
    ("Romil",                 "", "6/22/2026", "Rice (25 kg/sack)",  1, 1400, "CREDIT", ""),
    ("Jun2x B",               "", "6/22/2026", "Rice (25 kg/sack)",  1, 1400, "CREDIT", ""),
    ("Bryan",                 "", "6/22/2026", "Rice (25 kg/sack)",  1, 1400, "CREDIT", ""),
    ("JR",                    "", "6/22/2026", "Rice (25 kg/sack)",  1, 1400, "CREDIT", ""),
    ("Naldo",                 "", "6/22/2026", "Rice (25 kg/sack)",  1, 1400, "CREDIT", ""),
    ("Leonard",               "", "6/28/2026", "Rice (50 kg/sack)",  1, 2800, "CREDIT", "RED"),
    ("Jen2x Crossing Catmon", "", "7/1/2026",  "Rice (25 kg/sack)",  1, 1350, "CREDIT", ""),
    ("Tita Janeth",           "", "6/30/2026", "Rice (25 kg/sack)",  1, 1200, "CREDIT", ""),
    ("Julius",                "", "7/4/2026",  "Rice (25 kg/sack)",  1, 1400, "CREDIT", ""),
    ("Anghel",                "", "7/4/2026",  "Rice (25 kg/sack)",  1, 1400, "CREDIT", "RED"),
    ("Atan Verano",           "", "7/6/2026",  "Rice (50 kg/sack)",  1, 2500, "CREDIT", "RED"),
    ("Romel Sode",            "", "7/6/2026",  "Rice (25 kg/sack)",  1, 1400, "CREDIT", "RED"),
    ("Victor",                "", "7/6/2026",  "Rice (25 kg/sack)",  1, 1400, "CREDIT", "GREY"),
    ("Leni",                  "", "7/6/2026",  "Rice (25 kg/sack)",  1, 1400, "CREDIT", ""),
    ("Gulayan Nanay",         "", "7/7/2026",  "Rice (50 kg/sack)",  5, 2600, "CREDIT", ""),
    ("Papa",                  "", "7/11/2026", "Rice (25 kg/sack)",  1, 1200, "CREDIT", ""),
    ("Jun2x SM",              "", "7/12/2026", "Rice (25 kg/sack)",  1, 1400, "CREDIT", ""),
    ("Kadok",                 "", "7/12/2026", "Rice (25 kg/sack)",  1, 1400, "CREDIT", "RED"),
    ("Jun2x B",               "", "8/11/2026", "Rice (25 kg/sack)",  1, 1450, "CREDIT", ""),
    ("Bryan",                 "", "7/12/2026", "Rice (25 kg/sack)",  1, 1450, "CREDIT", ""),
    ("Jun2x SM",              "", "8/3/2026",  "Rice (25 kg/sack)",  1, 1450, "CREDIT", ""),
    ("Gulayan Nanay 2",       "", "8/3/2026",  "Rice (25 kg/sack)", 10, 1300, "CREDIT", ""),
    ("Julius",                "", "8/4/2026",  "Rice (25 kg/sack)",  1, 1450, "CREDIT", ""),
    ("Te lenie",              "", "8/5/2026",  "Rice (25 kg/sack)",  2, 1450, "CREDIT", ""),
    ("Romil",                 "", "8/5/2026",  "Rice (25 kg/sack)",  1, 1450, "CREDIT", ""),
    ("Gulayan Nanay 2",       "", "8/11/2026", "Rice (25 kg/sack)", 10, 1300, "CREDIT", "DUP?"),
    ("(name blank in source)","", "8/8/2026",  "Rice (25 kg/sack)",  1, 1450, "CREDIT", "NONAME"),
    ("Ranny",                 "", "8/18/2026", "Rice (25 kg/sack)",  1, 1450, "CREDIT", ""),
    ("Leonard",               "", "8/18/2026", "Rice (25 kg/sack)",  1, 1450, "CREDIT", ""),
    ("Anghel",                "", "8/18/2026", "Rice (25 kg/sack)",  1, 1450, "CREDIT", ""),
]

NOTE = {
    "RED":    "Highlighted red in the source ledger — meaning to be confirmed.",
    "GREY":   "Highlighted grey in the source ledger — meaning to be confirmed.",
    "DUP?":   "Same customer, qty and amount as an earlier row — confirm it is not a duplicate entry.",
    "NONAME": "Customer name was blank in the source ledger — please fill in.",
}

def d(s):
    m, day, y = s.split("/")
    return dt.date(int(y), int(m), int(day))

credit = [r for r in ROWS if r[6] == "CREDIT"]
cash   = [r for r in ROWS if r[6] != "CREDIT"]

PAY_FIRST, PAY_LAST = 6, 305          # payment entry rows
LED_FIRST = 6
LED_LAST  = LED_FIRST + len(credit) - 1
TOT_ROW   = LED_LAST + 2

wb = Workbook()

# ----------------------------------------------------------------- Credit Ledger
ws = wb.active
ws.title = "Credit Ledger"

def title_block(sh, title, sub, legend, width_last="M"):
    sh["A1"] = title
    sh["A1"].font = Font(name=FONT, size=16, bold=True, color=NAVY)
    sh["A2"] = sub
    sh["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")
    sh["A3"] = legend
    sh["A3"].font = Font(name=FONT, size=9, color="7F6000")
    sh["A3"].fill = PatternFill("solid", fgColor="FFF2CC")
    sh.merge_cells("A1:%s1" % width_last)
    sh.merge_cells("A2:%s2" % width_last)
    sh.merge_cells("A3:%s3" % width_last)
    sh.row_dimensions[1].height = 22
    sh.row_dimensions[3].height = 16

title_block(ws, "RICE CREDIT LEDGER — SALES ON CREDIT & OUTSTANDING BALANCES",
    "Source: 'Customer Sales & Loan Ledger' provided by the owner, 29 Aug 2026. "
    "Credit transactions only (%d of %d rows); the %d cash rows are on the 'Excluded (Cash)' sheet."
    % (len(credit), len(ROWS), len(cash)),
    "HOW TO USE: do not type anything on this sheet — it is all formulas. "
    "Record every payment on the 'Payments' sheet (yellow cells) using the Ledger ID from column A. "
    "Total Paid, Balance, Status and Last Payment update by themselves.")

HEAD = ["Ledger ID", "Customer Name", "Contact Number", "Transaction Date", "Product Type",
        "Qty (Sacks)", "Unit Price (PHP)", "Total Amount (PHP)", "Total Paid (PHP)",
        "Balance Due (PHP)", "Status", "Last Payment", "Notes"]
WIDTH = [9, 24, 16, 15, 19, 11, 15, 17, 15, 16, 11, 13, 46]

for i, h in enumerate(HEAD, start=1):
    c = ws.cell(row=5, column=i, value=h)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=HEADBG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = box
    ws.column_dimensions[get_column_letter(i)].width = WIDTH[i-1]
ws.row_dimensions[5].height = 30

PB = "Payments!$B$%d:$B$%d" % (PAY_FIRST, PAY_LAST)
PD = "Payments!$D$%d:$D$%d" % (PAY_FIRST, PAY_LAST)
PA = "Payments!$A$%d:$A$%d" % (PAY_FIRST, PAY_LAST)

for n, (cust, contact, date, prod, qty, price, method, flag) in enumerate(credit):
    r = LED_FIRST + n
    ws.cell(r, 1, n + 1)
    ws.cell(r, 2, cust)
    ws.cell(r, 3, contact)
    ws.cell(r, 4, d(date))
    ws.cell(r, 5, prod)
    ws.cell(r, 6, qty)
    ws.cell(r, 7, price)
    ws.cell(r, 8, "=F{0}*G{0}".format(r))
    ws.cell(r, 9, '=SUMIF({0},$A{1},{2})'.format(PB, r, PD))
    ws.cell(r, 10, "=H{0}-I{0}".format(r))
    ws.cell(r, 11, '=IF($H{0}=0,"",IF(ROUND($J{0},2)<0,"OVERPAID",'
                   'IF(ROUND($J{0},2)=0,"PAID",IF($I{0}>0,"PARTIAL","UNPAID"))))'.format(r))
    ws.cell(r, 12, '=IF(SUMPRODUCT(MAX(({1}=$A{2})*{0}))=0,"",'
                   'SUMPRODUCT(MAX(({1}=$A{2})*{0})))'.format(PA, PB, r))
    ws.cell(r, 13, NOTE.get(flag, ""))

    for col in range(1, 14):
        c = ws.cell(r, col)
        c.font = Font(name=FONT, size=10)
        c.border = box
        if n % 2 == 1:
            c.fill = PatternFill("solid", fgColor=BAND)
    ws.cell(r, 1).alignment = Alignment(horizontal="center")
    ws.cell(r, 4).number_format = "mm/dd/yyyy"
    ws.cell(r, 6).alignment = Alignment(horizontal="center")
    for col in (7, 8, 9, 10):
        ws.cell(r, col).number_format = "#,##0.00;(#,##0.00);-"
    ws.cell(r, 11).alignment = Alignment(horizontal="center")
    ws.cell(r, 11).font = Font(name=FONT, size=10, bold=True)
    ws.cell(r, 12).number_format = "mm/dd/yyyy"
    ws.cell(r, 12).alignment = Alignment(horizontal="center")
    ws.cell(r, 13).font = Font(name=FONT, size=9, italic=True, color="7F6000")
    if flag in ("RED", "GREY", "DUP?", "NONAME"):
        fill = PatternFill("solid", fgColor="FFC7CE" if flag == "RED" else "FFF2CC")
        for col in range(1, 14):
            ws.cell(r, col).fill = fill

ws.cell(TOT_ROW, 2, "TOTAL — %d credit sales" % len(credit))
ws.cell(TOT_ROW, 6, "=SUM(F{0}:F{1})".format(LED_FIRST, LED_LAST))
ws.cell(TOT_ROW, 8, "=SUM(H{0}:H{1})".format(LED_FIRST, LED_LAST))
ws.cell(TOT_ROW, 9, "=SUM(I{0}:I{1})".format(LED_FIRST, LED_LAST))
ws.cell(TOT_ROW, 10, "=SUM(J{0}:J{1})".format(LED_FIRST, LED_LAST))
for col in range(1, 14):
    c = ws.cell(TOT_ROW, col)
    c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.border = box
for col in (7, 8, 9, 10):
    ws.cell(TOT_ROW, col).number_format = "#,##0.00;(#,##0.00);-"
ws.cell(TOT_ROW, 6).alignment = Alignment(horizontal="center")

k = TOT_ROW + 2
ws.cell(k, 2, "Legend").font = Font(name=FONT, size=10, bold=True, color=NAVY)
legend = [
    ("Red rows", "Highlighted red in the original ledger. Copied across as-is — tell me what red means and I will turn it into a real column."),
    ("Yellow rows", "Something to confirm: grey highlight in the original, a possible duplicate, or a missing customer name."),
    ("Total Amount", "Qty x Unit Price. Recomputed from the source figures; every row matched the original ledger."),
    ("Total Paid", "The sum of every payment on the 'Payments' sheet carrying this Ledger ID."),
    ("Status", "PAID when the balance reaches zero, PARTIAL after a part payment, UNPAID until the first payment, OVERPAID if the payments exceed the sale."),
]
for i, (a, b) in enumerate(legend):
    ws.cell(k + 1 + i, 2, a).font = Font(name=FONT, size=9, bold=True)
    ws.cell(k + 1 + i, 3, b).font = Font(name=FONT, size=9, color="595959")

ws.freeze_panes = "C6"
ws.auto_filter.ref = "A5:M%d" % LED_LAST
ws.sheet_view.showGridLines = False

# ----------------------------------------------------------------- Payments
ps = wb.create_sheet("Payments")
title_block(ps, "PAYMENTS RECEIVED", 
    "One row per payment. A customer may pay in as many instalments as they like — just add another row.",
    "FILL IN THE YELLOW COLUMNS ONLY: Payment Date, Ledger ID, Amount Paid, Method, Notes. "
    "The Customer column fills itself from the Ledger ID, and the ledger updates automatically.", "F")

PH = ["Payment Date", "Ledger ID", "Customer (auto)", "Amount Paid (PHP)", "Method", "Notes"]
PW = [15, 11, 26, 18, 14, 40]
for i, h in enumerate(PH, start=1):
    c = ps.cell(row=5, column=i, value=h)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=HEADBG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = box
    ps.column_dimensions[get_column_letter(i)].width = PW[i-1]
ps.row_dimensions[5].height = 30

for r in range(PAY_FIRST, PAY_LAST + 1):
    ps.cell(r, 3, '=IF($B{0}="","",IFERROR(INDEX(\'Credit Ledger\'!$B${1}:$B${2},'
                  'MATCH($B{0},\'Credit Ledger\'!$A${1}:$A${2},0)),"** ID not found **"))'
                  .format(r, LED_FIRST, LED_LAST))
    for col in range(1, 7):
        c = ps.cell(r, col)
        c.font = Font(name=FONT, size=10)
        c.border = box
        if col in (1, 2, 4, 5, 6):
            c.fill = PatternFill("solid", fgColor=YELLOW)
    ps.cell(r, 1).number_format = "mm/dd/yyyy"
    ps.cell(r, 2).alignment = Alignment(horizontal="center")
    ps.cell(r, 3).font = Font(name=FONT, size=10, italic=True, color="595959")
    ps.cell(r, 4).number_format = "#,##0.00;(#,##0.00);-"

ptot = PAY_LAST + 2
ps.cell(ptot, 3, "TOTAL COLLECTED")
ps.cell(ptot, 4, "=SUM(D{0}:D{1})".format(PAY_FIRST, PAY_LAST))
for col in range(1, 7):
    c = ps.cell(ptot, col)
    c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.border = box
ps.cell(ptot, 4).number_format = "#,##0.00;(#,##0.00);-"

# example block, off to the side so it is never counted in any total
ps["H5"] = "EXAMPLE — reference only, nothing here is counted"
ps["H5"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
ps.merge_cells("H5:M5")
for i, h in enumerate(PH, start=8):
    c = ps.cell(6, i, h)
    c.font = Font(name=FONT, size=9, bold=True, color="595959")
    c.fill = PatternFill("solid", fgColor=GREY)
    c.border = box
    ps.column_dimensions[get_column_letter(i)].width = 16
_i = next(n for n, x in enumerate(credit) if x[0] == "Anghel" and x[2] == "7/4/2026")
EX_ID, EX_AMT = _i + 1, credit[_i][4] * credit[_i][5]
ex = [dt.date(2026, 8, 30), EX_ID, "Anghel", EX_AMT / 2,
      "part payment, balance %s" % format(EX_AMT / 2, ",.2f")]
ex.insert(4, "CASH")
for i, v in enumerate(ex, start=8):
    c = ps.cell(7, i, v)
    c.font = Font(name=FONT, size=9, italic=True, color="595959")
    c.border = box
ps.cell(7, 8).number_format = "mm/dd/yyyy"
ps.cell(7, 11).number_format = "#,##0.00"
ps["H9"] = ("Anghel's 04 Jul sale is Ledger ID %d for %s. Typing that row on the left makes the "
            "ledger show Total Paid %s, Balance %s, Status PARTIAL. Pay the rest later and it turns PAID."
            % (EX_ID, format(EX_AMT, ",.2f"), format(EX_AMT / 2, ",.2f"), format(EX_AMT / 2, ",.2f")))
ps["H9"].font = Font(name=FONT, size=9, color="595959")
ps.merge_cells("H9:M10")
ps["H9"].alignment = Alignment(wrap_text=True, vertical="top")

dv_id = DataValidation(type="whole", operator="between", formula1=1, formula2=len(credit),
                       allow_blank=True, showErrorMessage=True,
                       errorTitle="Unknown Ledger ID",
                       error="Enter a Ledger ID between 1 and %d, from column A of the Credit Ledger sheet." % len(credit))
ps.add_data_validation(dv_id)
dv_id.add("B%d:B%d" % (PAY_FIRST, PAY_LAST))

dv_m = DataValidation(type="list", formula1='"CASH,GCASH,BANK TRANSFER,RICE RETURNED,OTHER"',
                      allow_blank=True, showErrorMessage=False)
ps.add_data_validation(dv_m)
dv_m.add("E%d:E%d" % (PAY_FIRST, PAY_LAST))

dv_amt = DataValidation(type="decimal", operator="greaterThan", formula1=0, allow_blank=True,
                        showErrorMessage=True, errorTitle="Amount must be positive",
                        error="Enter the amount received, greater than zero.")
ps.add_data_validation(dv_amt)
dv_amt.add("D%d:D%d" % (PAY_FIRST, PAY_LAST))

ps.freeze_panes = "A6"
ps.sheet_view.showGridLines = False

# ----------------------------------------------------------------- Customer Summary
cs = wb.create_sheet("Customer Summary")
title_block(cs, "OUTSTANDING BALANCE BY CUSTOMER",
    "Every figure here is a formula over the Credit Ledger. Nothing to type.",
    "Use this sheet when collecting: it is the running utang of each customer across all their sales.", "G")

CH = ["Customer Name", "Sales", "Sacks", "Total Credit (PHP)", "Total Paid (PHP)", "Balance Due (PHP)", "Status"]
CW = [26, 9, 9, 18, 17, 18, 12]
for i, h in enumerate(CH, start=1):
    c = cs.cell(row=5, column=i, value=h)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=HEADBG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = box
    cs.column_dimensions[get_column_letter(i)].width = CW[i-1]
cs.row_dimensions[5].height = 30

names = sorted({r[0] for r in credit}, key=lambda s: s.lower())
LB = "'Credit Ledger'!$B${0}:$B${1}".format(LED_FIRST, LED_LAST)
for n, name in enumerate(names):
    r = 6 + n
    cs.cell(r, 1, name)
    cs.cell(r, 2, '=COUNTIF({0},$A{1})'.format(LB, r))
    cs.cell(r, 3, "=SUMIF({0},$A{1},'Credit Ledger'!$F${2}:$F${3})".format(LB, r, LED_FIRST, LED_LAST))
    cs.cell(r, 4, "=SUMIF({0},$A{1},'Credit Ledger'!$H${2}:$H${3})".format(LB, r, LED_FIRST, LED_LAST))
    cs.cell(r, 5, "=SUMIF({0},$A{1},'Credit Ledger'!$I${2}:$I${3})".format(LB, r, LED_FIRST, LED_LAST))
    cs.cell(r, 6, "=D{0}-E{0}".format(r))
    cs.cell(r, 7, '=IF(ROUND(F{0},2)<0,"OVERPAID",IF(ROUND(F{0},2)=0,"PAID",'
                  'IF(E{0}>0,"PARTIAL","UNPAID")))'.format(r))
    for col in range(1, 8):
        c = cs.cell(r, col)
        c.font = Font(name=FONT, size=10)
        c.border = box
        if n % 2 == 1:
            c.fill = PatternFill("solid", fgColor=BAND)
    for col in (2, 3, 7):
        cs.cell(r, col).alignment = Alignment(horizontal="center")
    for col in (4, 5, 6):
        cs.cell(r, col).number_format = "#,##0.00;(#,##0.00);-"
    cs.cell(r, 7).font = Font(name=FONT, size=10, bold=True)

ctot = 6 + len(names) + 1
cs.cell(ctot, 1, "TOTAL — %d customers" % len(names))
for col, letter in ((2, "B"), (3, "C"), (4, "D"), (5, "E"), (6, "F")):
    cs.cell(ctot, col, "=SUM({0}6:{0}{1})".format(letter, ctot - 2))
for col in range(1, 8):
    c = cs.cell(ctot, col)
    c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.border = box
for col in (4, 5, 6):
    cs.cell(ctot, col).number_format = "#,##0.00;(#,##0.00);-"
for col in (2, 3):
    cs.cell(ctot, col).alignment = Alignment(horizontal="center")

cs.cell(ctot + 2, 1, "Note: names are grouped exactly as written in the source ledger. "
                     "If two spellings are the same person (e.g. 'Ate Lenie' / 'Te lenie'), tell me and I will merge them.").font = Font(name=FONT, size=9, italic=True, color="7F6000")
cs.freeze_panes = "A6"
cs.auto_filter.ref = "A5:G%d" % (ctot - 2)
cs.sheet_view.showGridLines = False

# ----------------------------------------------------------------- Excluded (Cash)
xs = wb.create_sheet("Excluded (Cash)")
title_block(xs, "ROWS EXCLUDED FROM THE CREDIT LEDGER",
    "These two rows are marked CASH in the source ledger, so they carry no balance and are kept here for the record.",
    "If either one was actually on credit, tell me and I will move it into the Credit Ledger.", "G")
XH = ["Customer Name", "Transaction Date", "Product Type", "Qty (Sacks)", "Unit Price (PHP)", "Total Amount (PHP)", "Payment Method"]
for i, h in enumerate(XH, start=1):
    c = xs.cell(row=5, column=i, value=h)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=HEADBG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = box
    xs.column_dimensions[get_column_letter(i)].width = [24, 16, 19, 12, 16, 18, 16][i-1]
xs.row_dimensions[5].height = 30
for n, (cust, contact, date, prod, qty, price, method, flag) in enumerate(cash):
    r = 6 + n
    xs.cell(r, 1, cust); xs.cell(r, 2, d(date)); xs.cell(r, 3, prod)
    xs.cell(r, 4, qty); xs.cell(r, 5, price)
    xs.cell(r, 6, "=D{0}*E{0}".format(r)); xs.cell(r, 7, method)
    for col in range(1, 8):
        c = xs.cell(r, col); c.font = Font(name=FONT, size=10); c.border = box
    xs.cell(r, 2).number_format = "mm/dd/yyyy"
    xs.cell(r, 4).alignment = Alignment(horizontal="center")
    xs.cell(r, 7).alignment = Alignment(horizontal="center")
    for col in (5, 6):
        xs.cell(r, col).number_format = "#,##0.00;(#,##0.00);-"
xs.sheet_view.showGridLines = False

ws["A1"].comment = Comment(
    "Transcribed from the owner's 'Customer Sales & Loan Ledger' screenshot, 29 Aug 2026. "
    "Unit prices, quantities and dates are exactly as written there; Total Amount is recomputed "
    "as Qty x Unit Price and agreed with every figure in the source.", "Claude", height=140, width=320)

wb.save(OUT)
print("saved", OUT, "ledger rows", len(credit), "customers", len(names))
